from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Alignment
import anthropic
import base64
import io
import os
import json

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "CDF_Template_Base.xlsx")
SHEET_NAME    = "CDF Template FR to EN"

ITEM_DICT = """Jus → Juice
Paudre chocolat → Cocoa powder
Pomme de terre → Potato
Quaker oats, 500g → Quaker oats, 500g
Frommage → Cheese
Lait Inyange → Milk, Inyange brand
Yaourt → Yogurt
Banane plantain → Banana plantains
Fruits mixte → Fruits
Beure d'arachide → Peanut Butter
Chocolate Crème → Chocolate cream
Condiment → Condiments, assorted
Levire → Yeast
Miel, American Green, 1kg → Honey, American Green, 1kg
Ndizi, viazi, mambo → Potato, banana, yam
Ovaltine → Ovaltine
Sucre → Sugar
Manioc → Cassava
Blando de poulet → Chicken Breast
Oeuf Local, 10ct → Local eggs, 10 count
Oeufs, 30ct → Eggs, 30 count
Poisson filet → Fish fillet
poisson fimee → Smoked fish
Poisson Frais → Fresh fish
Poulet → Chicken, whole
Poulet fillet → Chicken, filet
Poulet village → Local chicken, whole
Viande de Boeuf → Beef
Viande de Echevre → Goat
Viande de Mouton → Lamb
Viande de Pors → Pork
Viande Hache → Ground meat
Champignon → Mushroom
Legume mixte → Vegetables
Lentil → Lentils
Tomatte Frais → Fresh tomato"""


def get_client():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY is not set")
    try:
        return anthropic.Anthropic(api_key=api_key)
    except Exception as e:
        raise ValueError(f"Failed to initialise Anthropic client: {str(e)}")


def fmt_currency(currency):
    return '#,##0.00" CDF"' if currency == "CDF" else '"$"#,##0.00'


def set_num(ws, addr, value, num_format):
    """Write a numeric value with centre alignment and number format."""
    cell = ws[addr]
    cell.value = value
    cell.number_format = num_format
    existing = cell.alignment
    cell.alignment = Alignment(
        horizontal="center",
        vertical=existing.vertical or "center",
        wrap_text=existing.wrap_text
    )


def safe_fill(template_path, fill_fn):
    """Load template, fill it, return bytes. Logo has been removed from template."""
    wb = openpyxl.load_workbook(template_path)
    fill_fn(wb)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "api_key_set": bool(os.environ.get("ANTHROPIC_API_KEY"))})


# -- Scan handwritten request sheet ------------------------------------------
@app.route("/scan-request", methods=["POST", "OPTIONS"])
def scan_request():
    if request.method == "OPTIONS":
        return "", 204

    try:
        client = get_client()

        if "image" not in request.files:
            return jsonify({"error": "No image uploaded"}), 400

        file       = request.files["image"]
        image_data = file.read()
        if not image_data:
            return jsonify({"error": "Image is empty"}), 400

        media_type = file.content_type or "image/jpeg"
        if media_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
            media_type = "image/jpeg"

        b64 = base64.standard_b64encode(image_data).decode("utf-8")

        prompt = f"""You are reading a handwritten purchase request sheet written in French for an NGO in the Democratic Republic of Congo.

Extract every line item visible on the sheet. For each item return:
- description_fr: the item description exactly as written in French
- description_en: the English translation (use the dictionary below if the item matches, otherwise translate accurately)
- unit: unit of measure (e.g. kg, pcs, litre, sachet, boite, divers, régime, bassin) — copy exactly as written
- qty: quantity as a number
- unit_price: unit price as a number (whole number only — no currency symbols, no commas, no "fc" suffix)
- cout_total: the line total as written in the "Coût Total" column (whole number, strip "fc" and commas — same rules as unit_price)

Also return one additional object at the END of the array for the grand total row:
{{"is_grand_total": true, "grand_total": <number as written in the Total row at the bottom, strip "fc" and commas>}}

CRITICAL RULES FOR READING UNIT PRICES AND TOTALS:
All prices are in Congolese Francs (CDF). Key facts about how they are written:

1. PRICES END IN "fc" — prices are written as "22,000fc" or "22.000fc". Strip the "fc" suffix and read only the number before it.
2. COMMAS ARE THOUSAND SEPARATORS — "22,000" means twenty-two thousand (22000). Never return 22 or 220 or 2200 for this.
3. ALL PRICES ARE WHOLE NUMBERS — there are no decimal points. If you see what looks like a decimal, it is a thousand separator.
4. EXPECTED RANGE — unit prices will almost always be between 1,000 and 200,000 CDF. If you read a price outside this range, re-examine it.
5. PRICES TYPICALLY END IN 000 — most prices are round thousands: 5,000 / 15,000 / 22,000 / 50,000 / 96,000. If you read a price that does not end in 000, double-check it.
6. DIGIT CONFUSION — watch carefully for these pairs which are frequently confused in this handwriting style: 0 vs 6, 1 vs 7, 3 vs 8, 9 vs 4, 5 vs 6, 2 vs 9. When two readings are plausible, prefer the one that gives a rounder, more common CDF price (e.g. 50,000 is more common than 60,000; 26,000 is more common than 96,000).
7. ROW DATES — some rows contain only a date (e.g. "Vendredi 27.03.2026"). These are section headers, not line items. Skip them entirely.
8. TOTAL ROW — read the grand total value at the bottom and return it as the final object with is_grand_total:true.

KNOWN ITEM DICTIONARY (French → English):
{ITEM_DICT}

For items not in the dictionary, provide an accurate English translation.
If a value is truly illegible after careful examination, use null for that field.
Return ONLY a valid JSON array, no markdown, no explanation.
Example: [{{"description_fr":"Viande de boeuf","description_en":"Beef","unit":"kg","qty":12,"unit_price":22000,"cout_total":264000}},{{"is_grand_total":true,"grand_total":1938000}}]"""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                    {"type": "text", "text": prompt}
                ]
            }]
        )

        text  = "".join(b.text for b in message.content if hasattr(b, "text"))
        clean = text.replace("```json", "").replace("```", "").strip()
        raw   = json.loads(clean)

        # Separate line items from the grand total object
        items       = [r for r in raw if not r.get("is_grand_total")]
        grand_total_obj = next((r for r in raw if r.get("is_grand_total")), None)
        grand_total = grand_total_obj.get("grand_total") if grand_total_obj else None

        return jsonify({"items": items, "grand_total": grand_total})

    except ValueError as e:
        return jsonify({"error": str(e)}), 503
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Could not parse AI response: {str(e)}"}), 422
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# -- Fill CDF template --------------------------------------------------------
@app.route("/fill-cdf-base", methods=["POST", "OPTIONS"])
def fill_cdf_base():
    if request.method == "OPTIONS":
        return "", 204

    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data received"}), 400

        requestor      = data.get("requestor", "")
        location       = data.get("location", "")
        date_submitted = data.get("date_submitted", "")
        speedkey       = data.get("speedkey", "")
        account_no     = data.get("account_no", "750300")
        currency       = data.get("currency", "CDF")
        items          = data.get("items", [])
        curr_fmt       = fmt_currency(currency)

        def fill(wb):
            ws = wb[SHEET_NAME]

            # Header
            ws["C4"] = requestor
            ws["I5"] = location
            ws["L2"] = date_submitted
            ws["L3"] = date_submitted
            ws["H6"] = "CDF" if currency == "CDF" else "USD"
            ws["K6"] = "X" if currency == "CDF" else ""
            ws["I6"] = "X" if currency == "USD" else ""

            # Line items (rows 22–46, max 25)
            # Col A row numbers: A22=1 (hardcoded), A23–A46 are =A22+1 etc — DO NOT overwrite
            # Col J Est Price: formula =IF(G22="","",G22*H22) — DO NOT overwrite
            # Col L Actual Price: leave BLANK — DO NOT write anything
            # J47 Total: formula =SUM($J$22:$K$46) — DO NOT overwrite
            for i, item in enumerate(items):
                if i >= 25:
                    break
                row        = 22 + i
                desc_fr    = item.get("description_fr") or ""
                desc_en    = item.get("description_en") or ""
                unit       = item.get("unit") or ""
                qty        = float(item.get("qty") or 0)
                unit_price = float(item.get("unit_price") or 0)

                ws[f"B{row}"].value     = desc_fr
                ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                ws[f"C{row}"].value     = desc_en
                ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                ws[f"D{row}"].value     = unit
                ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Speedkey
                ws[f"E{row}"].value     = speedkey
                ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Account number
                ws[f"F{row}"].value     = account_no
                ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Qty — number format, centred
                set_num(ws, f"G{row}", qty, "0")

                # Unit price — currency format, centred
                set_num(ws, f"H{row}", unit_price, curr_fmt)

                # J (Est Price) has =IF(G22="","",G22*H22) formula — leave it alone
                # L (Actual Price) — leave blank

            # Apply currency format to J column totals row so it renders correctly
            ws["J47"].number_format = curr_fmt

            # Submitted by
            ws["C70"] = requestor

        result = safe_fill(TEMPLATE_PATH, fill)

        safe_name = requestor.replace(" ", "_") or "Staff"
        date_str  = date_submitted.replace("/", "").replace("-", "") or "nodate"
        filename  = f"CDF_Base_{safe_name}_{date_str}.xlsx"

        return send_file(
            io.BytesIO(result),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
